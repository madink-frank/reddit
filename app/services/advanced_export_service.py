"""
Advanced Export Service - Enhanced multi-format export with charts and visualizations
"""

import asyncio
import tempfile
import os
from typing import List, Dict, Any, Optional, Union
from datetime import datetime, timedelta
import json
import csv
import io

from sqlalchemy.orm import Session
from fastapi import HTTPException

from app.models.post import Post
from app.models.user_billing import PointTransaction
from app.schemas.analytics import ExportRequest


class AdvancedExportService:
    """Enhanced export service with advanced formatting and visualization capabilities"""
    
    def __init__(self, db: Session):
        self.db = db
    
    async def create_excel_with_charts(
        self, 
        data: List[Dict[str, Any]], 
        title: str = "Data Export",
        include_charts: bool = True,
        include_pivot: bool = True
    ) -> str:
        """Create Excel file with advanced formatting, charts, and pivot tables"""
        
        try:
            import pandas as pd
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import LineChart, BarChart, PieChart, ScatterChart, Reference
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.worksheet.table import Table, TableStyleInfo
            import matplotlib.pyplot as plt
            import seaborn as sns
        except ImportError:
            raise HTTPException(
                status_code=500, 
                detail="Advanced Excel export requires pandas, openpyxl, matplotlib, seaborn"
            )
        
        if not data:
            raise ValueError("No data provided for export")
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # 1. Raw Data Sheet
        ws_data = wb.create_sheet("Raw Data")
        
        # Add data with table formatting
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_data.append(r)
        
        # Create Excel table
        table = Table(displayName="DataTable", ref=f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}{len(df)+1}")
        style = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=True
        )
        table.tableStyleInfo = style
        ws_data.add_table(table)
        
        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column_letter].width = adjusted_width
        
        # 2. Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, df, title)
        
        # 3. Charts Sheet
        if include_charts:
            ws_charts = wb.create_sheet("Charts")
            await self._create_excel_charts(ws_charts, df, ws_data)
        
        # 4. Pivot Tables Sheet
        if include_pivot:
            ws_pivot = wb.create_sheet("Pivot Analysis")
            self._create_pivot_analysis(ws_pivot, df)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
    
    def _create_summary_sheet(self, worksheet, df: 'pd.DataFrame', title: str):
        """Create summary statistics sheet"""
        
        # Title
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=18, bold=True)
        worksheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        worksheet['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        
        # Basic info
        info_data = [
            ["Report Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")],
            ["Total Records", len(df)],
            ["Total Columns", len(df.columns)],
            ["Memory Usage", f"{df.memory_usage(deep=True).sum() / 1024:.1f} KB"],
        ]
        
        row = 3
        for label, value in info_data:
            worksheet[f'A{row}'] = label
            worksheet[f'B{row}'] = value
            worksheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Column information
        row += 2
        worksheet[f'A{row}'] = "Column Information"
        worksheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        # Headers
        headers = ["Column", "Type", "Non-Null Count", "Unique Values", "Sample Values"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row += 1
        
        # Column details
        for col_name in df.columns:
            col_data = df[col_name]
            sample_values = ", ".join(str(x) for x in col_data.dropna().unique()[:3])
            
            worksheet.cell(row=row, column=1, value=col_name)
            worksheet.cell(row=row, column=2, value=str(col_data.dtype))
            worksheet.cell(row=row, column=3, value=col_data.count())
            worksheet.cell(row=row, column=4, value=col_data.nunique())
            worksheet.cell(row=row, column=5, value=sample_values[:50] + "..." if len(sample_values) > 50 else sample_values)
            row += 1
        
        # Numeric statistics
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            row += 2
            worksheet[f'A{row}'] = "Numeric Statistics"
            worksheet[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            stats_headers = ["Column", "Mean", "Median", "Std Dev", "Min", "Max"]
            for col, header in enumerate(stats_headers, 1):
                cell = worksheet.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            row += 1
            
            for col_name in numeric_cols:
                col_data = df[col_name]
                worksheet.cell(row=row, column=1, value=col_name)
                worksheet.cell(row=row, column=2, value=round(col_data.mean(), 2))
                worksheet.cell(row=row, column=3, value=round(col_data.median(), 2))
                worksheet.cell(row=row, column=4, value=round(col_data.std(), 2))
                worksheet.cell(row=row, column=5, value=round(col_data.min(), 2))
                worksheet.cell(row=row, column=6, value=round(col_data.max(), 2))
                row += 1
    
    async def _create_excel_charts(self, worksheet, df: 'pd.DataFrame', data_sheet):
        """Create various charts in Excel"""
        
        chart_row = 1
        
        # 1. Numeric distribution charts
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            for i, col in enumerate(numeric_cols[:3]):  # Limit to 3 charts
                chart = LineChart()
                chart.title = f"{col} Distribution"
                chart.style = 13
                chart.height = 10
                chart.width = 15
                
                # Add data reference
                data_ref = Reference(
                    data_sheet, 
                    min_col=list(df.columns).index(col) + 1,
                    min_row=2,
                    max_row=min(100, len(df) + 1)
                )
                chart.add_data(data_ref, titles_from_data=False)
                
                worksheet.add_chart(chart, f"A{chart_row}")
                chart_row += 15
        
        # 2. Categorical distribution
        categorical_cols = df.select_dtypes(include=['object']).columns
        if len(categorical_cols) > 0:
            col = categorical_cols[0]
            value_counts = df[col].value_counts().head(10)
            
            # Create bar chart
            chart = BarChart()
            chart.title = f"Top {col} Distribution"
            chart.style = 10
            chart.height = 10
            chart.width = 15
            
            worksheet.add_chart(chart, f"A{chart_row}")
    
    def _create_pivot_analysis(self, worksheet, df: 'pd.DataFrame'):
        """Create pivot table analysis"""
        
        worksheet['A1'] = "Pivot Analysis"
        worksheet['A1'].font = Font(size=16, bold=True)
        
        # Basic pivot analysis
        numeric_cols = df.select_dtypes(include=['number']).columns
        categorical_cols = df.select_dtypes(include=['object']).columns
        
        if len(numeric_cols) > 0 and len(categorical_cols) > 0:
            # Create simple pivot table
            try:
                pivot = df.groupby(categorical_cols[0])[numeric_cols[0]].agg(['count', 'mean', 'sum']).reset_index()
                
                # Add pivot data to worksheet
                row = 3
                headers = [categorical_cols[0], 'Count', 'Average', 'Total']
                for col, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                
                row += 1
                for _, pivot_row in pivot.iterrows():
                    worksheet.cell(row=row, column=1, value=pivot_row[categorical_cols[0]])
                    worksheet.cell(row=row, column=2, value=pivot_row['count'])
                    worksheet.cell(row=row, column=3, value=round(pivot_row['mean'], 2))
                    worksheet.cell(row=row, column=4, value=round(pivot_row['sum'], 2))
                    row += 1
                    
            except Exception as e:
                worksheet['A3'] = f"Pivot analysis failed: {str(e)}"
    
    async def create_advanced_pdf_report(
        self,
        data: List[Dict[str, Any]],
        title: str = "Advanced Data Report",
        include_visualizations: bool = True,
        template: str = "standard"
    ) -> str:
        """Create advanced PDF report with multiple visualizations and professional formatting"""
        
        try:
            import pandas as pd
            import matplotlib.pyplot as plt
            import seaborn as sns
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from io import BytesIO
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="Advanced PDF export requires pandas, matplotlib, seaborn, reportlab"
            )
        
        if not data:
            raise ValueError("No data provided for export")
        
        df = pd.DataFrame(data)
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
        
        # Set up matplotlib for high-quality charts
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")
        plt.rcParams['figure.dpi'] = 300
        plt.rcParams['savefig.dpi'] = 300
        
        # Create PDF document
        doc = SimpleDocTemplate(temp_file.name, pagesize=A4, topMargin=1*inch)
        styles = getSampleStyleSheet()
        story = []
        
        # Custom styles based on template
        if template == "executive":
            title_color = colors.HexColor('#1f4e79')
            accent_color = colors.HexColor('#2e86ab')
        elif template == "detailed":
            title_color = colors.HexColor('#2d5016')
            accent_color = colors.HexColor('#4a7c59')
        else:  # standard
            title_color = colors.HexColor('#2E86AB')
            accent_color = colors.HexColor('#A23B72')
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,
            textColor=title_color
        )
        
        # Build report content
        story.extend(await self._create_pdf_title_page(df, title, title_style, styles))
        
        if template == "executive":
            story.extend(await self._create_executive_summary(df, styles, accent_color))
        
        story.extend(await self._create_data_overview(df, styles, accent_color))
        
        if include_visualizations:
            chart_paths = await self._create_pdf_visualizations(df)
            story.extend(await self._add_visualizations_to_pdf(chart_paths, styles, accent_color))
        
        if template == "detailed":
            story.extend(await self._create_detailed_analysis(df, styles, accent_color))
        
        story.extend(await self._create_data_sample_table(df, styles, accent_color))
        
        # Build PDF
        doc.build(story)
        temp_file.close()
        
        return temp_file.name
    
    async def _create_pdf_title_page(self, df, title, title_style, styles):
        """Create PDF title page"""
        story = []
        
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 50))
        
        # Report metadata
        metadata = f"""
        <b>Generated:</b> {datetime.utcnow().strftime('%B %d, %Y at %H:%M UTC')}<br/>
        <b>Records:</b> {len(df):,}<br/>
        <b>Columns:</b> {len(df.columns)}<br/>
        <b>Data Size:</b> {df.memory_usage(deep=True).sum() / 1024:.1f} KB<br/>
        """
        
        story.append(Paragraph(metadata, styles['Normal']))
        story.append(PageBreak())
        
        return story
    
    async def _create_executive_summary(self, df, styles, accent_color):
        """Create executive summary section"""
        story = []
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            textColor=accent_color
        )
        
        story.append(Paragraph("Executive Summary", subtitle_style))
        
        # Key insights
        numeric_cols = df.select_dtypes(include=['number']).columns
        categorical_cols = df.select_dtypes(include=['object']).columns
        
        insights = []
        
        if len(numeric_cols) > 0:
            top_numeric = numeric_cols[0]
            insights.append(f"• Average {top_numeric}: {df[top_numeric].mean():.2f}")
            insights.append(f"• Total {top_numeric}: {df[top_numeric].sum():,.0f}")
        
        if len(categorical_cols) > 0:
            top_category = df[categorical_cols[0]].value_counts().index[0]
            count = df[categorical_cols[0]].value_counts().iloc[0]
            insights.append(f"• Most common {categorical_cols[0]}: {top_category} ({count} occurrences)")
        
        insights.append(f"• Data completeness: {(df.count().sum() / (len(df) * len(df.columns)) * 100):.1f}%")
        
        summary_text = "<br/>".join(insights)
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        return story
    
    async def _create_data_overview(self, df, styles, accent_color):
        """Create data overview section"""
        story = []
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            textColor=accent_color
        )
        
        story.append(Paragraph("Data Overview", subtitle_style))
        
        # Column information table
        col_info = []
        col_info.append(['Column', 'Type', 'Non-Null', 'Unique', 'Sample'])
        
        for col in df.columns[:10]:  # Limit to 10 columns
            col_data = df[col]
            sample = str(col_data.dropna().iloc[0]) if len(col_data.dropna()) > 0 else "N/A"
            col_info.append([
                col,
                str(col_data.dtype),
                str(col_data.count()),
                str(col_data.nunique()),
                sample[:20] + "..." if len(sample) > 20 else sample
            ])
        
        table = Table(col_info)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), accent_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8FF')])
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
        
        return story
    
    async def _create_pdf_visualizations(self, df):
        """Create visualizations for PDF"""
        chart_paths = []
        
        numeric_cols = df.select_dtypes(include=['number']).columns
        categorical_cols = df.select_dtypes(include=['object']).columns
        date_cols = df.select_dtypes(include=['datetime64']).columns
        
        # 1. Correlation heatmap for numeric data
        if len(numeric_cols) > 1:
            fig, ax = plt.subplots(figsize=(10, 8))
            correlation_matrix = df[numeric_cols].corr()
            sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', center=0, ax=ax)
            ax.set_title('Correlation Matrix', fontsize=16, fontweight='bold')
            
            chart_path = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            plt.savefig(chart_path.name, dpi=300, bbox_inches='tight')
            chart_paths.append(chart_path.name)
            plt.close()
        
        # 2. Distribution plots
        if len(numeric_cols) > 0:
            fig, axes = plt.subplots(2, 2, figsize=(12, 10))
            axes = axes.flatten()
            
            for i, col in enumerate(numeric_cols[:4]):
                ax = axes[i]
                df[col].hist(bins=20, ax=ax, alpha=0.7, edgecolor='black')
                ax.set_title(f'Distribution of {col}')
                ax.set_xlabel(col)
                ax.set_ylabel('Frequency')
                ax.grid(True, alpha=0.3)
            
            # Hide unused subplots
            for i in range(len(numeric_cols), 4):
                axes[i].set_visible(False)
            
            plt.tight_layout()
            
            chart_path = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            plt.savefig(chart_path.name, dpi=300, bbox_inches='tight')
            chart_paths.append(chart_path.name)
            plt.close()
        
        # 3. Categorical analysis
        if len(categorical_cols) > 0:
            fig, ax = plt.subplots(figsize=(12, 6))
            
            top_categories = df[categorical_cols[0]].value_counts().head(10)
            bars = ax.bar(range(len(top_categories)), top_categories.values)
            
            ax.set_title(f'Top {categorical_cols[0]} Distribution', fontsize=14, fontweight='bold')
            ax.set_xlabel(categorical_cols[0])
            ax.set_ylabel('Count')
            ax.set_xticks(range(len(top_categories)))
            ax.set_xticklabels(top_categories.index, rotation=45, ha='right')
            
            # Add value labels
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom')
            
            plt.tight_layout()
            
            chart_path = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            plt.savefig(chart_path.name, dpi=300, bbox_inches='tight')
            chart_paths.append(chart_path.name)
            plt.close()
        
        # 4. Time series if date columns exist
        if len(date_cols) > 0 and len(numeric_cols) > 0:
            fig, ax = plt.subplots(figsize=(12, 6))
            
            df_sorted = df.sort_values(date_cols[0])
            ax.plot(df_sorted[date_cols[0]], df_sorted[numeric_cols[0]], 
                   marker='o', linewidth=2, markersize=3)
            
            ax.set_title(f'{numeric_cols[0]} Over Time', fontsize=14, fontweight='bold')
            ax.set_xlabel(date_cols[0])
            ax.set_ylabel(numeric_cols[0])
            ax.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            
            plt.tight_layout()
            
            chart_path = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            plt.savefig(chart_path.name, dpi=300, bbox_inches='tight')
            chart_paths.append(chart_path.name)
            plt.close()
        
        return chart_paths
    
    async def _add_visualizations_to_pdf(self, chart_paths, styles, accent_color):
        """Add visualizations to PDF story"""
        story = []
        
        if not chart_paths:
            return story
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            textColor=accent_color
        )
        
        story.append(PageBreak())
        story.append(Paragraph("Data Visualizations", subtitle_style))
        
        for chart_path in chart_paths:
            try:
                img = Image(chart_path, width=7*inch, height=5.25*inch)
                story.append(img)
                story.append(Spacer(1, 20))
                
                # Clean up
                os.unlink(chart_path)
            except Exception as e:
                print(f"Failed to add chart: {e}")
        
        return story
    
    async def _create_detailed_analysis(self, df, styles, accent_color):
        """Create detailed statistical analysis"""
        story = []
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            textColor=accent_color
        )
        
        story.append(PageBreak())
        story.append(Paragraph("Detailed Statistical Analysis", subtitle_style))
        
        numeric_cols = df.select_dtypes(include=['number']).columns
        
        if len(numeric_cols) > 0:
            # Statistical summary table
            stats_data = [['Metric'] + list(numeric_cols[:5])]  # Limit columns
            
            metrics = ['Count', 'Mean', 'Std', 'Min', '25%', '50%', '75%', 'Max']
            desc = df[numeric_cols[:5]].describe()
            
            for metric in metrics:
                row = [metric]
                for col in numeric_cols[:5]:
                    if metric in desc.index:
                        value = desc.loc[metric, col]
                        row.append(f"{value:.2f}" if isinstance(value, (int, float)) else str(value))
                    else:
                        row.append("N/A")
                stats_data.append(row)
            
            table = Table(stats_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), accent_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8FF')])
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
        
        return story
    
    async def _create_data_sample_table(self, df, styles, accent_color):
        """Create data sample table"""
        story = []
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            textColor=accent_color
        )
        
        story.append(PageBreak())
        story.append(Paragraph("Data Sample", subtitle_style))
        
        # Limit data for display
        display_df = df.head(20)
        display_cols = df.columns[:6]  # Limit columns
        
        table_data = [list(display_cols)]
        for _, row in display_df.iterrows():
            table_data.append([
                str(row[col])[:25] + ('...' if len(str(row[col])) > 25 else '') 
                for col in display_cols
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), accent_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F8FF')])
        ]))
        
        story.append(table)
        
        return story
    
    async def create_csv_with_custom_fields(
        self,
        data: List[Dict[str, Any]],
        selected_fields: Optional[List[str]] = None,
        delimiter: str = ',',
        include_headers: bool = True
    ) -> str:
        """Create CSV with customizable field selection"""
        
        if not data:
            raise ValueError("No data provided for export")
        
        df = pd.DataFrame(data)
        
        # Select fields
        if selected_fields:
            available_fields = [field for field in selected_fields if field in df.columns]
            df = df[available_fields]
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        
        # Write CSV
        df.to_csv(
            temp_file.name,
            sep=delimiter,
            index=False,
            header=include_headers,
            encoding='utf-8'
        )
        
        temp_file.close()
        return temp_file.name
    
    async def estimate_export_resources(self, request: ExportRequest, data_count: int) -> Dict[str, Any]:
        """Estimate resources needed for export"""
        
        base_processing_time = 5  # seconds
        
        # Time estimates based on format and data size
        format_multipliers = {
            'csv': 1.0,
            'json': 1.2,
            'excel': 3.0,
            'pdf': 5.0,
            'xml': 2.0
        }
        
        # Size estimates (bytes per record)
        size_per_record = {
            'csv': 200,
            'json': 300,
            'excel': 500,
            'pdf': 1000,
            'xml': 400
        }
        
        format_multiplier = format_multipliers.get(request.format, 1.0)
        estimated_time = base_processing_time + (data_count / 1000) * format_multiplier
        
        estimated_size = data_count * size_per_record.get(request.format, 300)
        
        # Point cost calculation
        base_cost = 5
        size_cost = max(1, data_count // 100)  # 1 point per 100 records
        format_cost = int(format_multiplier * 2)
        
        total_cost = base_cost + size_cost + format_cost
        
        return {
            "estimatedTime": estimated_time,
            "estimatedSize": estimated_size,
            "pointsCost": total_cost,
            "processingComplexity": "low" if data_count < 1000 else "medium" if data_count < 10000 else "high"
        }