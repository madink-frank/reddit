from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, Query
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
import asyncio
import json
import csv
import io
import os
import tempfile
from datetime import datetime, timedelta
import uuid

from app.core.dependencies import get_db, get_current_user
from app.models.user import User
from app.models.post import Post
from app.models.user_billing import UserBilling, PointTransaction
from app.schemas.analytics import ExportRequest, ExportResult
from app.services.billing_service import BillingService
from app.services.analytics_service import AnalyticsService

router = APIRouter()

# ============================================================================
# Export Management
# ============================================================================

# In-memory storage for export jobs (in production, use Redis or database)
export_jobs: Dict[str, Dict[str, Any]] = {}

@router.post("/create", response_model=ExportResult)
async def create_export(
    request: ExportRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new export request"""
    
    # Generate export ID
    export_id = str(uuid.uuid4())
    
    # Validate request and estimate costs
    validation = await validate_export_request_internal(request, db, current_user)
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail=validation["errors"])
    
    # Check if user has enough points
    billing_service = BillingService(db)
    user_billing = billing_service.get_user_billing(current_user.id)
    
    if user_billing.current_points < validation["pointsCost"]:
        raise HTTPException(
            status_code=402, 
            detail=f"Insufficient points. Required: {validation['pointsCost']}, Available: {user_billing.current_points}"
        )
    
    # Create export job
    export_job = {
        "id": export_id,
        "requestId": export_id,
        "status": "queued",
        "progress": 0,
        "userId": current_user.id,
        "request": request.dict(),
        "createdAt": datetime.utcnow(),
        "estimatedSize": validation.get("estimatedSize", 0),
        "estimatedTime": validation.get("estimatedTime", 0),
        "pointsConsumed": 0,
        "expiresAt": datetime.utcnow() + timedelta(days=7),
    }
    
    export_jobs[export_id] = export_job
    
    # Start background processing
    background_tasks.add_task(process_export, export_id, request, current_user.id, db)
    
    return ExportResult(
        id=export_id,
        requestId=export_id,
        status="queued",
        progress=0,
        pointsConsumed=0,
        expiresAt=export_job["expiresAt"]
    )

@router.get("/{export_id}/status", response_model=ExportResult)
async def get_export_status(
    export_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get export status and result"""
    
    if export_id not in export_jobs:
        raise HTTPException(status_code=404, detail="Export not found")
    
    job = export_jobs[export_id]
    
    # Check if user owns this export
    if job["userId"] != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return ExportResult(
        id=job["id"],
        requestId=job["requestId"],
        status=job["status"],
        progress=job.get("progress", 0),
        downloadUrl=job.get("downloadUrl"),
        fileSize=job.get("fileSize"),
        recordCount=job.get("recordCount"),
        processingTime=job.get("processingTime"),
        pointsConsumed=job["pointsConsumed"],
        expiresAt=job["expiresAt"],
        error=job.get("error")
    )

@router.get("/{export_id}/download")
async def download_export(
    export_id: str,
    current_user: User = Depends(get_current_user)
):
    """Download export file"""
    
    if export_id not in export_jobs:
        raise HTTPException(status_code=404, detail="Export not found")
    
    job = export_jobs[export_id]
    
    # Check if user owns this export
    if job["userId"] != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if job["status"] != "completed":
        raise HTTPException(status_code=400, detail="Export not ready for download")
    
    if not job.get("filePath") or not os.path.exists(job["filePath"]):
        raise HTTPException(status_code=404, detail="Export file not found")
    
    # Determine content type based on format
    format_type = job["request"]["format"]
    content_types = {
        "csv": "text/csv",
        "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "json": "application/json",
        "pdf": "application/pdf",
        "xml": "application/xml"
    }
    
    content_type = content_types.get(format_type, "application/octet-stream")
    filename = f"export_{export_id}.{format_type}"
    
    return FileResponse(
        job["filePath"],
        media_type=content_type,
        filename=filename
    )

@router.delete("/{export_id}/cancel")
async def cancel_export(
    export_id: str,
    current_user: User = Depends(get_current_user)
):
    """Cancel an ongoing export"""
    
    if export_id not in export_jobs:
        raise HTTPException(status_code=404, detail="Export not found")
    
    job = export_jobs[export_id]
    
    # Check if user owns this export
    if job["userId"] != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if job["status"] in ["completed", "failed"]:
        raise HTTPException(status_code=400, detail="Cannot cancel completed or failed export")
    
    job["status"] = "cancelled"
    job["error"] = "Export cancelled by user"
    
    return {"message": "Export cancelled successfully"}

@router.get("/history", response_model=List[ExportResult])
async def get_export_history(
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(get_current_user)
):
    """Get list of user's exports"""
    
    user_exports = [
        job for job in export_jobs.values() 
        if job["userId"] == current_user.id
    ]
    
    # Sort by creation date (newest first)
    user_exports.sort(key=lambda x: x["createdAt"], reverse=True)
    
    # Apply pagination
    paginated_exports = user_exports[offset:offset + limit]
    
    return [
        ExportResult(
            id=job["id"],
            requestId=job["requestId"],
            status=job["status"],
            progress=job.get("progress", 0),
            downloadUrl=job.get("downloadUrl"),
            fileSize=job.get("fileSize"),
            recordCount=job.get("recordCount"),
            processingTime=job.get("processingTime"),
            pointsConsumed=job["pointsConsumed"],
            expiresAt=job["expiresAt"],
            error=job.get("error")
        )
        for job in paginated_exports
    ]

# ============================================================================
# Export Processing Functions
# ============================================================================

async def process_export(export_id: str, request: ExportRequest, user_id: int, db: Session):
    """Background task to process export request"""
    
    job = export_jobs[export_id]
    
    try:
        job["status"] = "processing"
        job["startTime"] = datetime.utcnow()
        
        # Deduct points upfront
        billing_service = BillingService(db)
        estimated_cost = await estimate_export_cost(request, db)
        
        billing_service.deduct_points(
            user_id=user_id,
            amount=estimated_cost,
            operation="export",
            description=f"Export {request.dataType} as {request.format}"
        )
        
        job["pointsConsumed"] = estimated_cost
        
        # Process based on data type and format
        if request.dataType == "posts":
            result = await export_posts_data(request, db, job)
        elif request.dataType == "analysis":
            result = await export_analysis_data(request, db, job)
        elif request.dataType == "images":
            result = await export_image_data(request, db, job)
        elif request.dataType == "reports":
            result = await export_reports_data(request, db, job)
        elif request.dataType == "metrics":
            result = await export_metrics_data(request, db, job)
        else:
            raise ValueError(f"Unsupported data type: {request.dataType}")
        
        # Update job with results
        job.update(result)
        job["status"] = "completed"
        job["progress"] = 100
        job["processingTime"] = (datetime.utcnow() - job["startTime"]).total_seconds()
        
    except Exception as e:
        job["status"] = "failed"
        job["error"] = str(e)
        job["processingTime"] = (datetime.utcnow() - job["startTime"]).total_seconds()

async def export_posts_data(request: ExportRequest, db: Session, job: Dict[str, Any]) -> Dict[str, Any]:
    """Export posts data in specified format"""
    
    # Query posts based on filters
    query = db.query(Post)
    
    if request.filters:
        if request.filters.get("dateRange"):
            date_range = request.filters["dateRange"]
            query = query.filter(
                Post.created_at >= date_range["start"],
                Post.created_at <= date_range["end"]
            )
        
        if request.filters.get("keywords"):
            keywords = request.filters["keywords"]
            for keyword in keywords:
                query = query.filter(Post.title.contains(keyword) | Post.content.contains(keyword))
        
        if request.filters.get("subreddits"):
            subreddits = request.filters["subreddits"]
            query = query.filter(Post.subreddit.in_(subreddits))
    
    # Apply limit
    max_records = request.options.get("maxRecords", 10000) if request.options else 10000
    posts = query.limit(max_records).all()
    
    job["progress"] = 30
    job["recordCount"] = len(posts)
    
    # Convert to export format
    if request.format == "csv":
        return await export_to_csv(posts, job)
    elif request.format == "excel":
        return await export_to_excel(posts, job, include_charts=True)
    elif request.format == "json":
        return await export_to_json(posts, job)
    elif request.format == "pdf":
        return await export_to_pdf(posts, job, "Posts Report")
    else:
        raise ValueError(f"Unsupported format: {request.format}")

async def export_analysis_data(request: ExportRequest, db: Session, job: Dict[str, Any]) -> Dict[str, Any]:
    """Export analysis results data"""
    
    # This would query analysis results from database
    # For now, return mock data
    analysis_data = []
    
    job["progress"] = 50
    job["recordCount"] = len(analysis_data)
    
    if request.format == "excel":
        return await export_to_excel(analysis_data, job, include_charts=True)
    else:
        return await export_to_csv(analysis_data, job)

async def export_image_data(request: ExportRequest, db: Session, job: Dict[str, Any]) -> Dict[str, Any]:
    """Export image analysis data"""
    
    # This would query image analysis results from database
    image_data = []
    
    job["progress"] = 50
    job["recordCount"] = len(image_data)
    
    return await export_to_excel(image_data, job)

async def export_reports_data(request: ExportRequest, db: Session, job: Dict[str, Any]) -> Dict[str, Any]:
    """Export reports and billing data"""
    
    # Query billing data
    billing_data = []
    
    job["progress"] = 50
    job["recordCount"] = len(billing_data)
    
    if request.format == "pdf":
        return await export_to_pdf(billing_data, job, "Billing Report")
    else:
        return await export_to_excel(billing_data, job, include_charts=True)

async def export_metrics_data(request: ExportRequest, db: Session, job: Dict[str, Any]) -> Dict[str, Any]:
    """Export system metrics data"""
    
    metrics_data = []
    
    job["progress"] = 50
    job["recordCount"] = len(metrics_data)
    
    return await export_to_csv(metrics_data, job)

# ============================================================================
# Format-specific Export Functions
# ============================================================================

async def export_to_csv(data: List[Any], job: Dict[str, Any]) -> Dict[str, Any]:
    """Export data to CSV format"""
    
    if not data:
        raise ValueError("No data to export")
    
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False)
    
    try:
        writer = csv.DictWriter(temp_file, fieldnames=data[0].__dict__.keys())
        writer.writeheader()
        
        for item in data:
            if hasattr(item, '__dict__'):
                writer.writerow(item.__dict__)
            else:
                writer.writerow(item)
        
        temp_file.close()
        
        # Get file size
        file_size = os.path.getsize(temp_file.name)
        
        return {
            "filePath": temp_file.name,
            "fileSize": file_size,
            "downloadUrl": f"/api/v1/export/{job['id']}/download"
        }
        
    except Exception as e:
        temp_file.close()
        os.unlink(temp_file.name)
        raise e

async def export_to_excel(data: List[Any], job: Dict[str, Any], include_charts: bool = False) -> Dict[str, Any]:
    """Export data to Excel format with formatting and charts"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, BarChart, PieChart, Reference, ScatterChart
        from openpyxl.chart.axis import DateAxis
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        import matplotlib.pyplot as plt
        import seaborn as sns
        from io import BytesIO
        import base64
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export requires openpyxl, pandas, matplotlib packages")
    
    if not data:
        raise ValueError("No data to export")
    
    # Convert data to DataFrame for easier manipulation
    if hasattr(data[0], '__dict__'):
        df = pd.DataFrame([item.__dict__ for item in data])
    else:
        df = pd.DataFrame(data)
    
    # Create workbook with multiple sheets
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create main data sheet
    ws_data = wb.create_sheet("Data")
    
    # Add data to worksheet with formatting
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    
    # Format headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, len(df.columns) + 1):
        cell = ws_data.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Format data cells
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = ws_data.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[column_letter].width = adjusted_width
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Add summary statistics
    summary_data = [
        ["Summary Statistics", ""],
        ["Total Records", len(df)],
        ["Columns", len(df.columns)],
        ["Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")],
        ["", ""],
    ]
    
    # Add numeric column statistics
    numeric_cols = df.select_dtypes(include=['number']).columns
    if len(numeric_cols) > 0:
        summary_data.append(["Numeric Column Statistics", ""])
        for col in numeric_cols:
            summary_data.extend([
                [f"{col} - Mean", df[col].mean()],
                [f"{col} - Median", df[col].median()],
                [f"{col} - Std Dev", df[col].std()],
                [f"{col} - Min", df[col].min()],
                [f"{col} - Max", df[col].max()],
                ["", ""],
            ])
    
    for row_data in summary_data:
        ws_summary.append(row_data)
    
    # Format summary sheet
    for row in ws_summary.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and ("Statistics" in cell.value or "Summary" in cell.value):
                cell.font = Font(bold=True, size=14)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Add charts if requested and data is suitable
    if include_charts and len(df) > 1:
        try:
            ws_charts = wb.create_sheet("Charts")
            
            # Create charts based on data types
            chart_row = 1
            
            # 1. Bar chart for categorical data
            categorical_cols = df.select_dtypes(include=['object']).columns
            if len(categorical_cols) > 0 and len(numeric_cols) > 0:
                # Create a bar chart
                chart = BarChart()
                chart.title = f"Distribution by {categorical_cols[0]}"
                chart.style = 10
                chart.x_axis.title = categorical_cols[0]
                chart.y_axis.title = "Count"
                
                # Add data (simplified - would need proper aggregation in real implementation)
                data_ref = Reference(ws_data, min_col=1, min_row=1, max_row=min(20, len(df)+1), max_col=2)
                chart.add_data(data_ref, titles_from_data=True)
                
                ws_charts.add_chart(chart, f"A{chart_row}")
                chart_row += 15
            
            # 2. Line chart for time series data
            date_cols = df.select_dtypes(include=['datetime64']).columns
            if len(date_cols) > 0 and len(numeric_cols) > 0:
                chart = LineChart()
                chart.title = f"Trend Over Time"
                chart.style = 13
                chart.x_axis.title = "Date"
                chart.y_axis.title = numeric_cols[0] if len(numeric_cols) > 0 else "Value"
                
                # Add data
                data_ref = Reference(ws_data, min_col=1, min_row=1, max_row=min(100, len(df)+1), max_col=2)
                chart.add_data(data_ref, titles_from_data=True)
                
                ws_charts.add_chart(chart, f"A{chart_row}")
                chart_row += 15
            
            # 3. Pie chart for categorical distribution
            if len(categorical_cols) > 0:
                chart = PieChart()
                chart.title = f"Distribution of {categorical_cols[0]}"
                
                # Create aggregated data for pie chart (simplified)
                data_ref = Reference(ws_data, min_col=1, min_row=2, max_row=min(10, len(df)+1))
                labels_ref = Reference(ws_data, min_col=1, min_row=2, max_row=min(10, len(df)+1))
                
                chart.add_data(data_ref)
                chart.set_categories(labels_ref)
                
                ws_charts.add_chart(chart, f"H{chart_row}")
                
        except Exception as e:
            print(f"Chart creation failed: {e}")
            # Continue without charts if creation fails
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    temp_file.close()
    
    # Get file size
    file_size = os.path.getsize(temp_file.name)
    
    job["progress"] = 90
    
    return {
        "filePath": temp_file.name,
        "fileSize": file_size,
        "downloadUrl": f"/api/v1/export/{job['id']}/download"
    }

async def export_to_json(data: List[Any], job: Dict[str, Any]) -> Dict[str, Any]:
    """Export data to JSON format"""
    
    # Convert data to JSON-serializable format
    json_data = []
    for item in data:
        if hasattr(item, '__dict__'):
            json_data.append({k: str(v) if not isinstance(v, (str, int, float, bool, type(None))) else v 
                            for k, v in item.__dict__.items()})
        else:
            json_data.append(item)
    
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False)
    
    try:
        json.dump(json_data, temp_file, indent=2, default=str)
        temp_file.close()
        
        # Get file size
        file_size = os.path.getsize(temp_file.name)
        
        return {
            "filePath": temp_file.name,
            "fileSize": file_size,
            "downloadUrl": f"/api/v1/export/{job['id']}/download"
        }
        
    except Exception as e:
        temp_file.close()
        os.unlink(temp_file.name)
        raise e

async def export_to_pdf(data: List[Any], job: Dict[str, Any], title: str) -> Dict[str, Any]:
    """Export data to PDF format with visualizations"""
    
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.linecharts import HorizontalLineChart
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.charts.piecharts import Pie
        import matplotlib.pyplot as plt
        import seaborn as sns
        import pandas as pd
        from io import BytesIO
        import base64
    except ImportError:
        raise HTTPException(status_code=500, detail="PDF export requires reportlab, matplotlib, seaborn, pandas packages")
    
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
    
    try:
        # Convert data to DataFrame for analysis
        if hasattr(data[0], '__dict__'):
            df = pd.DataFrame([item.__dict__ for item in data])
        else:
            df = pd.DataFrame(data)
        
        # Create PDF document
        doc = SimpleDocTemplate(temp_file.name, pagesize=A4, topMargin=1*inch)
        styles = getSampleStyleSheet()
        story = []
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#2E86AB')
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            textColor=colors.HexColor('#A23B72')
        )
        
        # Add title page
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 30))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", subtitle_style))
        
        summary_text = f"""
        <b>Report Generated:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC<br/>
        <b>Total Records:</b> {len(data):,}<br/>
        <b>Data Columns:</b> {len(df.columns)}<br/>
        <b>Date Range:</b> {df.select_dtypes(include=['datetime64']).min().min() if len(df.select_dtypes(include=['datetime64']).columns) > 0 else 'N/A'} to {df.select_dtypes(include=['datetime64']).max().max() if len(df.select_dtypes(include=['datetime64']).columns) > 0 else 'N/A'}<br/>
        """
        
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Key Statistics
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            story.append(Paragraph("Key Statistics", subtitle_style))
            
            stats_data = [['Metric', 'Value']]
            for col in numeric_cols[:5]:  # Limit to first 5 numeric columns
                stats_data.extend([
                    [f'{col} - Average', f'{df[col].mean():.2f}'],
                    [f'{col} - Total', f'{df[col].sum():,.0f}'],
                    [f'{col} - Range', f'{df[col].min():.2f} - {df[col].max():.2f}']
                ])
            
            stats_table = Table(stats_data)
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F8FF')),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8FF')])
            ]))
            
            story.append(stats_table)
            story.append(Spacer(1, 20))
        
        # Generate and embed charts
        chart_images = []
        
        # Set matplotlib style for better-looking charts
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")
        
        # 1. Distribution chart for numeric data
        if len(numeric_cols) > 0:
            fig, ax = plt.subplots(figsize=(10, 6))
            df[numeric_cols[0]].hist(bins=20, ax=ax, alpha=0.7, color='skyblue', edgecolor='black')
            ax.set_title(f'Distribution of {numeric_cols[0]}', fontsize=14, fontweight='bold')
            ax.set_xlabel(numeric_cols[0])
            ax.set_ylabel('Frequency')
            ax.grid(True, alpha=0.3)
            
            # Save chart to bytes
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
            img_buffer.seek(0)
            
            # Create temporary image file
            chart_temp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            chart_temp.write(img_buffer.getvalue())
            chart_temp.close()
            chart_images.append(chart_temp.name)
            
            plt.close()
        
        # 2. Trend chart for time series data
        date_cols = df.select_dtypes(include=['datetime64']).columns
        if len(date_cols) > 0 and len(numeric_cols) > 0:
            fig, ax = plt.subplots(figsize=(12, 6))
            
            # Sort by date and plot
            df_sorted = df.sort_values(date_cols[0])
            ax.plot(df_sorted[date_cols[0]], df_sorted[numeric_cols[0]], 
                   marker='o', linewidth=2, markersize=4, alpha=0.8)
            
            ax.set_title(f'{numeric_cols[0]} Trend Over Time', fontsize=14, fontweight='bold')
            ax.set_xlabel(date_cols[0])
            ax.set_ylabel(numeric_cols[0])
            ax.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            
            # Save chart
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
            img_buffer.seek(0)
            
            chart_temp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            chart_temp.write(img_buffer.getvalue())
            chart_temp.close()
            chart_images.append(chart_temp.name)
            
            plt.close()
        
        # 3. Categorical distribution
        categorical_cols = df.select_dtypes(include=['object']).columns
        if len(categorical_cols) > 0:
            fig, ax = plt.subplots(figsize=(10, 6))
            
            # Get top categories
            top_categories = df[categorical_cols[0]].value_counts().head(10)
            
            bars = ax.bar(range(len(top_categories)), top_categories.values, 
                         color=sns.color_palette("husl", len(top_categories)))
            ax.set_title(f'Top {categorical_cols[0]} Distribution', fontsize=14, fontweight='bold')
            ax.set_xlabel(categorical_cols[0])
            ax.set_ylabel('Count')
            ax.set_xticks(range(len(top_categories)))
            ax.set_xticklabels(top_categories.index, rotation=45, ha='right')
            ax.grid(True, alpha=0.3, axis='y')
            
            # Add value labels on bars
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom')
            
            # Save chart
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
            img_buffer.seek(0)
            
            chart_temp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            chart_temp.write(img_buffer.getvalue())
            chart_temp.close()
            chart_images.append(chart_temp.name)
            
            plt.close()
        
        # Add charts to PDF
        if chart_images:
            story.append(PageBreak())
            story.append(Paragraph("Data Visualizations", subtitle_style))
            
            for i, chart_path in enumerate(chart_images):
                try:
                    # Add chart image
                    img = Image(chart_path, width=6*inch, height=3.6*inch)
                    story.append(img)
                    story.append(Spacer(1, 20))
                    
                    # Clean up temporary file
                    os.unlink(chart_path)
                except Exception as e:
                    print(f"Failed to add chart {i}: {e}")
        
        # Add data table (limited rows for PDF)
        story.append(PageBreak())
        story.append(Paragraph("Data Sample", subtitle_style))
        
        if len(df) > 0:
            # Limit columns and rows for readability
            display_df = df.head(50)  # First 50 rows
            display_cols = df.columns[:8]  # First 8 columns
            
            table_data = [list(display_cols)]
            for _, row in display_df.iterrows():
                table_data.append([str(row[col])[:30] + ('...' if len(str(row[col])) > 30 else '') 
                                 for col in display_cols])
            
            # Create table with better styling
            data_table = Table(table_data)
            data_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F8FF')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F8FF')])
            ]))
            
            story.append(data_table)
        
        # Add footer information
        story.append(Spacer(1, 30))
        footer_text = f"""
        <i>This report was generated automatically by the Reddit Content Platform.<br/>
        For questions or support, please contact the system administrator.<br/>
        Report ID: {job.get('id', 'N/A')}</i>
        """
        story.append(Paragraph(footer_text, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        temp_file.close()
        
        # Get file size
        file_size = os.path.getsize(temp_file.name)
        
        job["progress"] = 95
        
        return {
            "filePath": temp_file.name,
            "fileSize": file_size,
            "downloadUrl": f"/api/v1/export/{job['id']}/download"
        }
        
    except Exception as e:
        temp_file.close()
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        raise e

# ============================================================================
# Utility Functions
# ============================================================================

@router.post("/validate")
async def validate_export_request(
    request: ExportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Validate export request and return estimates"""
    return await validate_export_request_internal(request, db, current_user)

async def validate_export_request_internal(request: ExportRequest, db: Session, user: User) -> Dict[str, Any]:
    """Internal validation function"""
    
    errors = []
    warnings = []
    
    # Validate data type
    valid_data_types = ["posts", "analysis", "images", "reports", "metrics"]
    if request.dataType not in valid_data_types:
        errors.append(f"Invalid data type. Must be one of: {', '.join(valid_data_types)}")
    
    # Validate format
    valid_formats = ["csv", "excel", "json", "pdf", "xml"]
    if request.format not in valid_formats:
        errors.append(f"Invalid format. Must be one of: {', '.join(valid_formats)}")
    
    # Estimate size and cost
    estimated_records = await estimate_record_count(request, db)
    estimated_size = estimated_records * 1024  # Rough estimate: 1KB per record
    estimated_time = max(10, estimated_records / 100)  # Minimum 10 seconds, then 100 records/second
    points_cost = await estimate_export_cost(request, db)
    
    # Check limits
    max_records = request.options.get("maxRecords", 50000) if request.options else 50000
    if estimated_records > max_records:
        warnings.append(f"Export will be limited to {max_records} records (estimated: {estimated_records})")
    
    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "estimatedSize": estimated_size,
        "estimatedTime": estimated_time,
        "pointsCost": points_cost,
        "estimatedRecords": estimated_records
    }

async def estimate_record_count(request: ExportRequest, db: Session) -> int:
    """Estimate number of records for export"""
    
    if request.dataType == "posts":
        query = db.query(Post)
        
        if request.filters:
            if request.filters.get("dateRange"):
                date_range = request.filters["dateRange"]
                query = query.filter(
                    Post.created_at >= date_range["start"],
                    Post.created_at <= date_range["end"]
                )
            
            if request.filters.get("subreddits"):
                subreddits = request.filters["subreddits"]
                query = query.filter(Post.subreddit.in_(subreddits))
        
        return query.count()
    
    # Default estimates for other data types
    return 1000

async def estimate_export_cost(request: ExportRequest, db: Session) -> int:
    """Estimate point cost for export"""
    
    base_cost = 10  # Base cost for any export
    
    # Cost per record
    estimated_records = await estimate_record_count(request, db)
    record_cost = max(1, estimated_records // 100)  # 1 point per 100 records
    
    # Format multiplier
    format_multipliers = {
        "csv": 1.0,
        "json": 1.2,
        "excel": 2.0,
        "pdf": 3.0,
        "xml": 1.5
    }
    
    format_multiplier = format_multipliers.get(request.format, 1.0)
    
    total_cost = int((base_cost + record_cost) * format_multiplier)
    return max(5, total_cost)  # Minimum 5 points

@router.get("/templates")
async def get_export_templates():
    """Get available export templates"""
    
    templates = [
        {
            "id": "posts_excel",
            "name": "Posts Excel Report",
            "description": "Export posts with analysis data in Excel format with charts",
            "format": "excel",
            "dataType": "posts",
            "defaultOptions": {
                "includeAnalysis": True,
                "includeMetadata": True,
                "maxRecords": 10000
            }
        },
        {
            "id": "sentiment_pdf",
            "name": "Sentiment Analysis Report",
            "description": "PDF report with sentiment analysis and visualizations",
            "format": "pdf",
            "dataType": "analysis",
            "defaultOptions": {
                "includeAnalysis": True,
                "includeImages": True
            }
        },
        {
            "id": "billing_excel",
            "name": "Billing Report",
            "description": "Excel report with billing data and usage analytics",
            "format": "excel",
            "dataType": "reports",
            "defaultOptions": {
                "includeAnalysis": True,
                "includeMetadata": True
            }
        }
    ]
    
    return templates

@router.get("/fields/{data_type}")
async def get_available_fields(data_type: str):
    """Get available fields for CSV export"""
    
    field_mappings = {
        "posts": [
            {"field": "id", "label": "ID", "type": "string"},
            {"field": "title", "label": "Title", "type": "string"},
            {"field": "content", "label": "Content", "type": "string"},
            {"field": "author", "label": "Author", "type": "string"},
            {"field": "subreddit", "label": "Subreddit", "type": "string"},
            {"field": "score", "label": "Score", "type": "number"},
            {"field": "created_at", "label": "Created At", "type": "date"},
        ],
        "analysis": [
            {"field": "id", "label": "Analysis ID", "type": "string"},
            {"field": "text", "label": "Text", "type": "string"},
            {"field": "sentiment_score", "label": "Sentiment Score", "type": "number"},
            {"field": "sentiment_label", "label": "Sentiment Label", "type": "string"},
            {"field": "processed_at", "label": "Processed At", "type": "date"},
        ]
    }
    
    return field_mappings.get(data_type, [])

@router.get("/stats")
async def get_export_stats(current_user: User = Depends(get_current_user)):
    """Get export statistics"""
    
    user_exports = [job for job in export_jobs.values() if job["userId"] == current_user.id]
    
    total_exports = len(user_exports)
    successful_exports = len([job for job in user_exports if job["status"] == "completed"])
    success_rate = (successful_exports / total_exports * 100) if total_exports > 0 else 0
    
    # Calculate averages
    completed_exports = [job for job in user_exports if job["status"] == "completed"]
    avg_size = sum(job.get("fileSize", 0) for job in completed_exports) / len(completed_exports) if completed_exports else 0
    avg_time = sum(job.get("processingTime", 0) for job in completed_exports) / len(completed_exports) if completed_exports else 0
    
    # Popular formats
    format_counts = {}
    for job in user_exports:
        format_type = job["request"]["format"]
        format_counts[format_type] = format_counts.get(format_type, 0) + 1
    
    popular_formats = [{"format": k, "count": v} for k, v in format_counts.items()]
    popular_formats.sort(key=lambda x: x["count"], reverse=True)
    
    return {
        "totalExports": total_exports,
        "successRate": success_rate,
        "averageSize": avg_size,
        "averageTime": avg_time,
        "popularFormats": popular_formats,
        "recentActivity": []  # Would calculate from actual data
    }